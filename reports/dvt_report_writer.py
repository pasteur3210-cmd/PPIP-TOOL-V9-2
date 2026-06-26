from __future__ import annotations

from pathlib import Path
from datetime import date
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


FACTORY_TARGET_SHEETS = {"Monthly_Loading", "Weekly_Loading", "Date_Loading", "Risk_List", "已出貨_Shipped"}

def _prepare_output_df(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Apply user-facing output column rules."""
    if df.empty:
        return df
    df = df.copy()
    if sheet_name in FACTORY_TARGET_SHEETS:
        # The report should show Factory, not the old customer field.  Factory is
        # populated by the external factory_mapping.xlsx/csv file.  If no mapping
        # was provided or no rule matched, the value is Unknown and Check_Log records it.
        if "factory" in df.columns:
            if "customer" in df.columns:
                df = df.drop(columns=["customer"])
            df = df.rename(columns={"factory": "Factory"})
        elif "customer" in df.columns:
            df = df.rename(columns={"customer": "Factory"})
        if "Factory" in df.columns:
            cols = list(df.columns)
            cols.remove("Factory")
            insert_at = cols.index("region") + 1 if "region" in cols else 0
            cols.insert(insert_at, "Factory")
            df = df[cols]
    return df


def _write_df(ws, df: pd.DataFrame, title: str):
    """Write a titled dataframe with title row 1, header row 2, data from row 3."""
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    header_row = 2
    data_start_row = 3
    if df.empty:
        ws.cell(data_start_row, 1, "No data")
        return

    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_offset, row in enumerate(rows, start=header_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_offset, column=c_idx, value=value)

    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = f"A{data_start_row}"
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 35)



def _build_check_log(report_rows: list[dict], shipped_rows: list[dict], all_rows: list[dict], excluded_rows=None, sheet_mode: str = "first", exclude_green: bool = True) -> pd.DataFrame:
    """Build traceability log for selected sheets and shipped-green handling."""
    records = []
    excluded_rows = excluded_rows or []
    report_row_ids = {r.get("row_id") for r in report_rows}
    excluded_row_ids = {r.get("row_id") for r in excluded_rows}
    by_source_sheet: dict[tuple[str, str], dict] = {}
    if sheet_mode == "first":
        sheet_rule = "First Sheet only"
    elif sheet_mode == "manual":
        sheet_rule = "Manual selected sheets only"
    else:
        sheet_rule = "All valid sheets"
    green_rule = "green rows excluded from main report" if exclude_green else "green rows retained in main report"

    for r in all_rows:
        key = (str(r.get("source_file", "")), str(r.get("sheet_name", "")))
        item = by_source_sheet.setdefault(
            key,
            {
                "source_file": key[0],
                "selected_sheet": key[1],
                "sheet_mode": sheet_mode,
                "exclude_green_from_main_report": exclude_green,
                "total_rows_read": 0,
                "main_report_rows": 0,
                "shipped_green_rows": 0,
                "model_strikethrough_excluded_rows": 0,
                "factory_mapping_missing_rows": 0,
                "factory_mapping_matched_rows": 0,
                "factory_mapping_conflict_rows": 0,
                "factory_mapping_model_only_unique_rows": 0,
                "blank_current_fcd_rows_in_loading": 0,
                "blank_current_fcd_qty_in_loading": 0,
                "blank_current_fcd_included_in_loading": "YES",
                "total_shipment_qty": 0,
                "main_report_shipment_qty": 0,
                "shipped_shipment_qty": 0,
                "strikethrough_excluded_qty": 0,
                "rule": f"{sheet_rule}; green row rule=A~L >= 6 green cells; {green_rule}; green rows are always copied to 已出貨_Shipped; Model欄位刪除線一律排除並寫入排除資料_Excluded; Current FCD空白/TBD/非日期資料仍納入Loading並彙總到未排日期欄位",
            },
        )
        qty = int(r.get("shipment_qty") or 0)
        item["total_rows_read"] += 1
        item["total_shipment_qty"] += qty
        if r.get("is_shipped_green"):
            item["shipped_green_rows"] += 1
            item["shipped_shipment_qty"] += qty
        if r.get("row_id") in excluded_row_ids or r.get("is_model_strikethrough"):
            item["model_strikethrough_excluded_rows"] += 1
            item["strikethrough_excluded_qty"] += qty
        if (r.get("row_id") in report_row_ids and int(r.get("shipment_qty") or 0) > 0 and not isinstance(r.get("current_fcd"), date)):
            item["blank_current_fcd_rows_in_loading"] += 1
            item["blank_current_fcd_qty_in_loading"] += qty
        if r.get("factory_mapping_status") == "MISSING":
            item["factory_mapping_missing_rows"] += 1
        elif r.get("factory_mapping_status") == "CONFLICT":
            item["factory_mapping_conflict_rows"] += 1
        elif r.get("factory_mapping_status") == "MATCHED":
            item["factory_mapping_matched_rows"] += 1
            if r.get("factory_mapping_method") == "MODEL_ONLY_UNIQUE":
                item["factory_mapping_model_only_unique_rows"] += 1
        if r.get("row_id") in report_row_ids:
            item["main_report_rows"] += 1
            item["main_report_shipment_qty"] += qty
    records.extend(by_source_sheet.values())
    return pd.DataFrame(records)

def write_dvt_report(tables, rows, output_path, logger: logging.Logger | None = None, shipped_rows=None, all_rows=None, excluded_rows=None, sheet_mode: str = "first", exclude_green: bool = True):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if logger:
        logger.info("EXPORT_START | output=%s", output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_總覽"
    shipped_rows = shipped_rows or []
    excluded_rows = excluded_rows or []
    all_rows = all_rows or rows
    total_rows = len(all_rows)
    active_rows = len(rows)
    shipped_count = len(shipped_rows)
    excluded_count = len(excluded_rows)
    total_qty = sum(int(r.get("shipment_qty") or 0) for r in all_rows)
    active_qty = sum(int(r.get("shipment_qty") or 0) for r in rows)
    shipped_qty = sum(int(r.get("shipment_qty") or 0) for r in shipped_rows)
    excluded_qty = sum(int(r.get("shipment_qty") or 0) for r in excluded_rows)
    sched_qty = sum(int(r.get("shipment_qty") or 0) for r in rows if r.get("include_in_schedule"))
    blank_fcd_loading_qty = sum(int(r.get("shipment_qty") or 0) for r in rows if int(r.get("shipment_qty") or 0) > 0 and not isinstance(r.get("current_fcd"), date))
    loading_qty = sum(int(r.get("shipment_qty") or 0) for r in rows if int(r.get("shipment_qty") or 0) > 0)
    risk_rows = sum(1 for r in rows if r.get("risk_level") in ["MEDIUM", "HIGH"])
    ws.append(["PPIP DVT樣本格式生產排程報表"])
    ws.append(["KPI", "Value"])
    for row in [["Raw Rows", total_rows], ["Main Report Rows", active_rows], ["Shipped Green Rows", shipped_count], ["Model Strikethrough Excluded Rows", excluded_count], ["Raw Shipment Qty", total_qty], ["Main Report Qty", active_qty], ["Shipped Green Qty", shipped_qty], ["Model Strikethrough Excluded Qty", excluded_qty], ["Scheduled Qty (Dated FCD)", sched_qty], ["Blank Current FCD Loading Qty", blank_fcd_loading_qty], ["Total Loading Qty incl Blank FCD", loading_qty], ["Risk Rows", risk_rows]]:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=16)
    for c in ws[2]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    for name, key in [
        ("Monthly_Loading", "monthly"),
        ("Weekly_Loading", "weekly"),
        ("Date_Loading", "date_matrix"),
        ("Top20_Model_Loading", "top_model"),
        ("Risk_List", "risk"),
        ("Back log 明細", "raw"),
        ("已出貨_Shipped", "shipped"),
        ("排除資料_Excluded", "excluded"),
        ("Check_Log", "check_log"),
    ]:
        ws2 = wb.create_sheet(name[:31])
        if key == "raw":
            df = pd.DataFrame(rows)
        elif key == "shipped":
            df = pd.DataFrame(shipped_rows)
        elif key == "excluded":
            df = pd.DataFrame(excluded_rows)
        elif key == "check_log":
            df = _build_check_log(rows, shipped_rows, all_rows, excluded_rows=excluded_rows, sheet_mode=sheet_mode, exclude_green=exclude_green)
        else:
            df = pd.DataFrame(tables.get(key, []))
        df = _prepare_output_df(df, name)
        _write_df(ws2, df, name)
    wb.save(output_path)
    if logger:
        logger.info("EXPORT_SUCCESS | output=%s | sheets=%s", output_path, len(wb.sheetnames))
    return output_path
