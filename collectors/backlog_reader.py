from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile
import tempfile
import shutil
import time
import gc
import re
import pandas as pd
import logging
from openpyxl import load_workbook

REGION_KEYS = ["EU-Czech", "EU-Spain", "NA", "ROW"]


def _region_from_name(name: str) -> str:
    lowered = name.lower()
    for key in REGION_KEYS:
        if key.lower().replace("-", "") in lowered.replace("-", ""):
            return key
    if "czech" in lowered:
        return "EU-Czech"
    if "spain" in lowered:
        return "EU-Spain"
    if re.search(r"\bna\b", lowered):
        return "NA"
    if "row" in lowered:
        return "ROW"
    return "UNKNOWN"


def _standardize_columns(cols) -> list[str]:
    return [str(c).strip().replace("\n", " ") if str(c) != "nan" else "" for c in cols]


def _safe_rmtree(path: Path, logger: logging.Logger | None = None, retries: int = 5, delay: float = 0.5) -> None:
    """Remove temporary folder safely on Windows.

    Some Excel readers / antivirus / Windows indexing may keep a temporary xlsx file
    locked for a short time.  The old TemporaryDirectory cleanup raised WinError 32
    and made the whole report fail even after the data had been loaded.  This helper
    retries cleanup and finally ignores the cleanup error, because temp-file cleanup
    should not block report generation.
    """
    for attempt in range(1, retries + 1):
        try:
            gc.collect()
            shutil.rmtree(path)
            return
        except PermissionError as exc:
            if logger:
                logger.warning(
                    "TEMP_CLEANUP_RETRY | attempt=%s/%s | path=%s | error=%s",
                    attempt,
                    retries,
                    path,
                    exc,
                )
            time.sleep(delay)
        except FileNotFoundError:
            return
        except Exception as exc:
            if logger:
                logger.warning("TEMP_CLEANUP_WARNING | path=%s | error=%s", path, exc)
            return
    if logger:
        logger.warning("TEMP_CLEANUP_SKIPPED | path=%s | reason=file_still_locked", path)



def _excel_files_in_zip_temp(zip_path: Path, logger: logging.Logger | None = None) -> tuple[Path, list[Path]]:
    """Extract zip to a temp folder and return (tmpdir, excel_files). Caller cleans tmpdir."""
    tmpdir = Path(tempfile.mkdtemp(prefix="ppip_backlog_"))
    with ZipFile(zip_path) as zf:
        zf.extractall(tmpdir)
    files = [
        p
        for p in tmpdir.rglob("*")
        if p.suffix.lower() in [".xlsx", ".xls", ".xlsm"]
        and not p.name.startswith("~$")
        and "__MACOSX" not in p.parts
    ]
    if logger:
        logger.info("UNZIP_SOURCE | zip=%s | excel_files=%s", zip_path.name, len(files))
    return tmpdir, files


def _normalize_selected_sheets(selected_sheets) -> dict[str, set[str]]:
    """Normalize manual sheet selection into {filename: {sheet names}}."""
    result: dict[str, set[str]] = {}
    if not selected_sheets:
        return result
    for file_name, sheets in dict(selected_sheets).items():
        if sheets is None:
            continue
        if isinstance(sheets, str):
            sheets = [sheets]
        cleaned = {str(sheet) for sheet in sheets if str(sheet).strip()}
        if cleaned:
            result[str(file_name)] = cleaned
    return result


def _choose_sheets(path: Path, sheet_names: list[str], sheet_mode: str = "first", selected_sheets=None, logger: logging.Logger | None = None) -> list[str]:
    """Choose workbook sheets based on default-first, all, or manual selection."""
    if sheet_mode == "first":
        selected = sheet_names[:1]
        if logger and selected:
            logger.info("SHEET_MODE_FIRST | file=%s | selected_sheet=%s | total_sheets=%s", path.name, selected[0], len(sheet_names))
        return selected
    if sheet_mode == "manual":
        mapping = _normalize_selected_sheets(selected_sheets)
        wanted = mapping.get(path.name, set())
        selected = [name for name in sheet_names if name in wanted]
        if logger:
            logger.info("SHEET_MODE_MANUAL | file=%s | selected_sheets=%s | total_sheets=%s", path.name, selected, len(sheet_names))
        return selected
    selected = list(sheet_names)
    if logger:
        logger.info("SHEET_MODE_ALL | file=%s | total_sheets=%s", path.name, len(sheet_names))
    return selected


def inspect_backlog_zip_sheets(zip_path, logger: logging.Logger | None = None) -> list[dict]:
    """Return workbook/sheet metadata for GUI manual sheet selection.

    Output example:
    [{"source_file": "Forecast.xlsx", "region": "ROW", "sheets": ["Jun 18", "Jun 12"]}]
    """
    zip_path = Path(zip_path)
    if not zip_path.exists():
        raise FileNotFoundError(zip_path)
    tmpdir = None
    result: list[dict] = []
    try:
        tmpdir, files = _excel_files_in_zip_temp(zip_path, logger=logger)
        for file in sorted(files):
            sheets: list[str] = []
            try:
                if file.suffix.lower() in [".xlsx", ".xlsm"]:
                    wb = load_workbook(file, read_only=True, data_only=True)
                    try:
                        sheets = list(wb.sheetnames)
                    finally:
                        wb.close()
                else:
                    with pd.ExcelFile(file) as xls:
                        sheets = list(xls.sheet_names)
            except Exception as exc:
                if logger:
                    logger.exception("ERROR | INSPECT_WORKBOOK_FAILED | file=%s | error=%s", file.name, exc)
                sheets = []
            result.append({"source_file": file.name, "region": _region_from_name(file.name), "sheets": sheets})
    finally:
        if tmpdir is not None:
            _safe_rmtree(tmpdir, logger=logger)
    return result


def _cell_is_green(cell) -> bool:
    """Return True when a cell fill looks like the Excel shipped green marker.

    The customer marks shipped rows with bright green fill.  Excel files may store
    colors as ARGB/RGB, indexed, or theme colors, so we use a conservative rule:
    any explicit RGB color where green is dominant and high is treated as green.
    """
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type is None:
        return False
    color = fill.fgColor
    if color is None:
        return False
    rgb = None
    if getattr(color, "type", None) == "rgb" and color.rgb:
        rgb = str(color.rgb)
    elif getattr(color, "type", None) == "indexed":
        # Common Excel indexed bright green is 4.
        return color.indexed == 4
    if not rgb:
        return False
    rgb = rgb[-6:].upper()
    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
    except Exception:
        return False
    # Covers bright green such as 00FF00 / 92D050 while avoiding yellow/blue fills.
    return g >= 140 and g >= r + 45 and g >= b + 45


def _row_is_shipped_green(ws, excel_row: int, first_col: int = 1, last_col: int = 12) -> tuple[bool, int, int]:
    """Detect shipped rows by green fill count across key columns A~L."""
    cells = [ws.cell(excel_row, col) for col in range(first_col, min(last_col, ws.max_column) + 1)]
    checked = len(cells)
    green_count = sum(1 for cell in cells if _cell_is_green(cell))
    # User's shipped rows are painted green across the row. Use >= 6 key cells to
    # prevent one highlighted cell from excluding a normal backlog line.
    return green_count >= 6, green_count, checked



def _find_model_column(headers: list[str]) -> int | None:
    """Return 1-based Excel column index for the Model column."""
    for idx, header in enumerate(headers, start=1):
        text = str(header or "").strip().lower().replace("\n", " ")
        if text in {"model", "model name"} or text.startswith("model "):
            return idx
    for idx, header in enumerate(headers, start=1):
        if "model" in str(header or "").strip().lower():
            return idx
    return None


def _cell_has_strikethrough(cell) -> bool:
    """Return True when Excel font strike / strikethrough is enabled."""
    font = getattr(cell, "font", None)
    return bool(getattr(font, "strike", False))


def _row_has_model_strikethrough(ws, excel_row: int, model_col: int | None) -> bool:
    """Detect cancelled/invalid rows by Model cell strikethrough."""
    if not model_col:
        return False
    return _cell_has_strikethrough(ws.cell(excel_row, model_col))


def _read_excel_file_with_openpyxl(path: Path, region: str, logger: logging.Logger | None = None, sheet_mode: str = "first", selected_sheets=None) -> list[dict]:
    rows: list[dict] = []
    wb = None
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
        sheet_names = list(wb.sheetnames)
        if logger:
            logger.info("LOAD_WORKBOOK | file=%s | sheets=%s | engine=openpyxl", path.name, len(sheet_names))
        sheets_to_read = _choose_sheets(path, sheet_names, sheet_mode=sheet_mode, selected_sheets=selected_sheets, logger=logger)

        for sheet in sheets_to_read:
            ws = wb[sheet]
            if ws.max_row < 2:
                continue
            headers = _standardize_columns([ws.cell(1, c).value for c in range(1, ws.max_column + 1)])
            model_col = _find_model_column(headers)
            joined_cols = "|".join(headers).lower()
            if not any(k in joined_cols for k in ["model", "shipment", "current fcd", "sor"]):
                if logger:
                    logger.warning("SKIP_SHEET | file=%s | sheet=%s | reason=no_required_like_columns", path.name, sheet)
                continue
            loaded = 0
            shipped = 0
            strikethrough = 0
            for r in range(2, ws.max_row + 1):
                values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                if all(v is None or str(v).strip() == "" for v in values):
                    continue
                record = {headers[i] if headers[i] else f"Column_{i+1}": values[i] for i in range(len(headers))}
                is_shipped, green_count, checked_cells = _row_is_shipped_green(ws, r, 1, 12)
                is_strikethrough = _row_has_model_strikethrough(ws, r, model_col)
                record["__source_file"] = path.name
                record["__sheet_name"] = sheet
                record["__excel_row"] = r
                record["__is_shipped"] = is_shipped
                record["__is_strikethrough"] = is_strikethrough
                record["__strikethrough_checked_column"] = model_col or ""
                record["__green_cells"] = green_count
                record["__green_checked_cells"] = checked_cells
                record["region"] = region
                rows.append(record)
                loaded += 1
                if is_shipped:
                    shipped += 1
                if is_strikethrough:
                    strikethrough += 1
            if logger:
                logger.info("LOAD_SOURCE | file=%s | sheet=%s | region=%s | rows=%s | shipped_green_rows=%s | model_strikethrough_rows=%s | cols=%s", path.name, sheet, region, loaded, shipped, strikethrough, len(headers))
    finally:
        try:
            if wb is not None:
                wb.close()
        finally:
            gc.collect()
    return rows


def _read_excel_file(path: Path, region: str, logger: logging.Logger | None = None, sheet_mode: str = "first", selected_sheets=None) -> list[dict]:
    if path.suffix.lower() in [".xlsx", ".xlsm"]:
        return _read_excel_file_with_openpyxl(path, region, logger=logger, sheet_mode=sheet_mode, selected_sheets=selected_sheets)

    rows: list[dict] = []
    try:
        # Use a context manager so the workbook handle is closed before temp cleanup.
        with pd.ExcelFile(path) as xls:
            sheet_names = list(xls.sheet_names)
            if logger:
                logger.info("LOAD_WORKBOOK | file=%s | sheets=%s", path.name, len(sheet_names))
            sheets_to_read = _choose_sheets(path, sheet_names, sheet_mode=sheet_mode, selected_sheets=selected_sheets, logger=logger)

            for sheet in sheets_to_read:
                try:
                    # Read from the already-open ExcelFile object.  This prevents pandas
                    # from opening another handle to the same temporary xlsx file.
                    df = pd.read_excel(xls, sheet_name=sheet, dtype=object)
                    df.columns = _standardize_columns(df.columns)
                    df = df.dropna(how="all")
                    if df.empty:
                        continue
                    # Keep only sheets that look like forecast/backlog tables.
                    joined_cols = "|".join(df.columns).lower()
                    if not any(k in joined_cols for k in ["model", "shipment", "current fcd", "sor"]):
                        if logger:
                            logger.warning("SKIP_SHEET | file=%s | sheet=%s | reason=no_required_like_columns", path.name, sheet)
                        continue
                    df["__source_file"] = path.name
                    df["__sheet_name"] = sheet
                    df["region"] = region
                    recs = df.to_dict("records")
                    rows.extend(recs)
                    if logger:
                        logger.info(
                            "LOAD_SOURCE | file=%s | sheet=%s | region=%s | rows=%s | cols=%s",
                            path.name,
                            sheet,
                            region,
                            len(df),
                            len(df.columns),
                        )
                except Exception as exc:
                    if logger:
                        logger.exception("ERROR | LOAD_SHEET_FAILED | file=%s | sheet=%s | error=%s", path.name, sheet, exc)
                    raise
    except Exception as exc:
        if logger:
            logger.exception("ERROR | LOAD_WORKBOOK_FAILED | file=%s | error=%s", path, exc)
        raise
    finally:
        # Ensure openpyxl/pandas objects are released before deleting extracted files.
        gc.collect()
    return rows


def read_backlog_zip(zip_path, as_of_date=None, logger: logging.Logger | None = None, sheet_mode: str = "first", selected_sheets=None):
    """Read Forecast(order) backlog Excel files from a ZIP and return raw rows.

    Rules:
    - region is parsed from file name: EU-Czech/EU-Spain/NA/ROW.
    - by default only the left-most / first worksheet in each workbook is loaded to avoid historical sheets.
    - manual mode loads only sheets selected by the GUI/CLI.
    - every output row keeps __source_file, __sheet_name and region for traceability.
    """
    zip_path = Path(zip_path)
    if not zip_path.exists():
        if logger:
            logger.error("ERROR | INPUT_ZIP_NOT_FOUND | zip_path=%s", zip_path)
        raise FileNotFoundError(zip_path)

    raw_rows: list[dict] = []
    tmpdir = None
    try:
        tmpdir, files = _excel_files_in_zip_temp(zip_path, logger=logger)
        if logger:
            logger.info("UNZIP_SOURCE | zip=%s | excel_files=%s | as_of_date=%s | sheet_mode=%s", zip_path.name, len(files), as_of_date, sheet_mode)
        if not files:
            if logger:
                logger.error("ERROR | NO_EXCEL_FILES | zip=%s", zip_path)
            raise ValueError("ZIP內找不到Excel檔案")
        for file in sorted(files):
            region = _region_from_name(file.name)
            raw_rows.extend(_read_excel_file(file, region, logger=logger, sheet_mode=sheet_mode, selected_sheets=selected_sheets))
    finally:
        if tmpdir is not None:
            _safe_rmtree(tmpdir, logger=logger)

    if logger:
        logger.info("LOAD_SOURCE_DONE | total_raw_rows=%s", len(raw_rows))
    return raw_rows
